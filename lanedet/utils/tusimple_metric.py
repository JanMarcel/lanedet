import os
import numpy as np
import openpyxl.workbook
import openpyxl.worksheet
from sklearn.linear_model import LinearRegression
import json as json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.drawing.fill import PatternFillProperties

class LaneEval(object):
    lr = LinearRegression()
    pixel_thresh = 20
    pt_thresh = 0.85

    @staticmethod
    def get_angle(xs, y_samples):
        xs, ys = xs[xs >= 0], y_samples[xs >= 0]
        if len(xs) > 1:
            LaneEval.lr.fit(ys[:, None], xs)
            k = LaneEval.lr.coef_[0]
            theta = np.arctan(k)
        else:
            theta = 0
        return theta

    @staticmethod
    def line_accuracy(pred, gt, thresh):
        pred = np.array([p if p >= 0 else -100 for p in pred])
        gt = np.array([g if g >= 0 else -100 for g in gt])
        return np.sum(np.where(np.abs(pred - gt) < thresh, 1., 0.)) / len(gt)

    @staticmethod
    def bench(pred, gt, y_samples, running_time, sheet: openpyxl.worksheet = None):
        if any(len(p) != len(y_samples) for p in pred):
            raise Exception('Format of lanes error.')
        if running_time > 200 or len(gt) + 2 < len(pred):
            return 0., 0., 1.
        angles = [LaneEval.get_angle(
            np.array(x_gts), np.array(y_samples)) for x_gts in gt]
        threshs = [LaneEval.pixel_thresh / np.cos(angle) for angle in angles]
        line_accs = []
        fp, fn = 0., 0.
        matched = 0.
        for x_gts, thresh in zip(gt, threshs):
            accs = [LaneEval.line_accuracy(
                np.array(x_preds), np.array(x_gts), thresh) for x_preds in pred]
            max_acc = np.max(accs) if len(accs) > 0 else 0.
            if max_acc < LaneEval.pt_thresh:
                fn += 1
            else:
                matched += 1
            line_accs.append(max_acc)
        fp = len(pred) - matched
        if len(gt) > 4 and fn > 0:
            fn -= 1
        s = sum(line_accs)
        if len(gt) > 4:
            s -= min(line_accs)

        a = s / max(min(4.0, len(gt)), 1.)
        p = fp / len(pred) if len(pred) > 0 else 0.
        n = fn / max(min(len(gt), 4.), 1.)
        
        if sheet is not None:
            LaneEval.write_excel_sheet(sheet, a, p, n, pred, gt, y_samples)

        return a, p, n

    @staticmethod
    def bench_one_submit(pred_file, gt_file, create_excel: bool = True):
        try:
            json_pred = [json.loads(line)
                         for line in open(pred_file).readlines()]
        except BaseException as e:
            raise Exception('Fail to load json file of the prediction.')
        json_gt = [json.loads(line) for line in open(gt_file).readlines()]
        if len(json_gt) != len(json_pred):
            raise Exception(
                'We do not get the predictions of all the test tasks')
        gts = {l['raw_file']: l for l in json_gt}
        accuracy, fp, fn = 0., 0., 0.
        if create_excel:
            wb = openpyxl.Workbook()
        for pred in json_pred:
            if 'raw_file' not in pred or 'lanes' not in pred or 'run_time' not in pred:
                raise Exception(
                    'raw_file or lanes or run_time not in some predictions.')
            raw_file = pred['raw_file']
            pred_lanes = pred['lanes']
            run_time = pred['run_time']
            if raw_file not in gts:
                raise Exception(
                    'Some raw_file from your predictions do not exist in the test tasks.')

            gt = gts[raw_file]
            gt_lanes = gt['lanes']
            y_samples = gt['h_samples']
            try:
                sheet: openpyxl.worksheet = wb.create_sheet(os.path.basename(os.path.dirname(raw_file)))
                a, p, n = LaneEval.bench(
                    pred_lanes, gt_lanes, y_samples, run_time, sheet)

            except BaseException as e:
                raise Exception('Format of lanes error.')
            accuracy += a
            fp += p
            fn += n
                
        num = len(gts)
        
        if create_excel:
            wb.save(filename="evaluation.xlsx")

        # the first return parameter is the default ranking parameter
        return json.dumps([
            {'name': 'Accuracy', 'value': accuracy / num, 'order': 'desc'},
            {'name': 'FP', 'value': fp / num, 'order': 'asc'},
            {'name': 'FN', 'value': fn / num, 'order': 'asc'}
        ]), accuracy / num

    @staticmethod
    def write_excel_sheet(sheet: openpyxl.worksheet, accuracy: float, false_positive: float, false_negative: float, pred, gt, y_samples):
        sheet["A1"] = "Accuracy:"
        sheet["A2"] = "FalsePositive"
        sheet["A3"] = "FalseNegative"
        sheet["B1"] = accuracy
        sheet["B2"] = false_positive
        sheet["B3"] = false_negative

        sheet["A5"] = "y_samples"

        for i, y in enumerate(y_samples):
            sheet[f"A{i + 6}"] = y
        
        for i, gt_lane in enumerate(gt):
            sheet.cell(row=5, column=i + 2).value = "Label" + str(i)
            for j, x in enumerate(gt_lane):
                sheet.cell(row=j + 6, column=i + 2).value = x
        
        column_offset = len(gt) + 2
        for i, pred_lane in enumerate(pred):
            sheet.cell(row=5, column=i + column_offset).value = "Lane" + str(i)
            for j, x in enumerate(pred_lane):
                sheet.cell(row=j + 6, column=i + column_offset).value = x

if __name__ == '__main__':
    import sys
    try:
        if len(sys.argv) != 3:
            raise Exception('Invalid input arguments')
        print(LaneEval.bench_one_submit(sys.argv[1], sys.argv[2]))
    except Exception as e:
        print(e.message)
        sys.exit(e.message)
